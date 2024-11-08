import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from functions import select_excel_file, select_image_file, select_folder, find_url_column, edit_image, download_image, open_excel
from model import load_ai_model, predict_color

current_directory = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(current_directory, 'files', 'eticaretmerkezim_buyuk_logo.png')
ico_path = os.path.join(current_directory, 'files', 'eticaretmerkezim_logo.ico')

class Application:
    def __init__(self, root):
        self.root = root
        self.root.title("eticaretmerkezim AI")
        root.iconbitmap(ico_path)
        self.root.geometry("800x450")
        self.root.configure(bg="#B0B0B0")

        # Model ve dosya yolları
        self.model = load_ai_model()
        self.excel_file = None
        self.folder_of_path = None
        self.image_path = None
        self.ai_excel_file_path = None

        self.create_widgets()
        
        # Sağ alt köşe çerçevesi
        bottom_right_frame = tk.Frame(root, bg="#B0B0B0")
        bottom_right_frame.grid(row=2, column=1, sticky="se", padx=15, pady=5)
        root.grid_rowconfigure(2, weight=1)
        root.grid_columnconfigure(1, weight=1)

        # Logoyu yükleyin
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((180, 50), Image.LANCZOS)  # Logonun boyutunu ayarlayın
        logo_photo = ImageTk.PhotoImage(logo_image)

        # Logoyu ve sürüm bilgisini ekleyin
        logo_label = tk.Label(bottom_right_frame, image=logo_photo, bg="#B0B0B0")
        logo_label.image = logo_photo  # Referansı saklayın
        logo_label.pack(side="top", padx=(0, 17), pady=(0,5))  # Logoyu üstte yerleştirin

        # Sağ alt köşeye sürüm numarası ve program adını ekleyin
        version_label = tk.Label(self.root, text="v1.0", font=("Arial", 8), bg="#B0B0B0")
        version_label.grid(row=2, column=1, sticky="se", padx=(0,7), pady=(0,18))
        
        # Sağ alt köşeye yazılım bilgisi ekleyin
        info_label = tk.Label(self.root, text="Bu yazılım, Emircan Akyüz tarafından geliştirilmiştir. Tüm hakları saklıdır. © 2024", font=("Arial", 8), bg="#B0B0B0")
        info_label.grid(row=2, column=1, sticky="se", padx=(0,7), pady=(0,2))

    def create_widgets(self):
        # Ana frame'lerin büyümeye duyarlı olması için ayarlar
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)

        # Excel Frame
        excel_frame = tk.Frame(self.root, bg="#B0B0B0")
        excel_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        excel_frame.columnconfigure(0, weight=1)
        excel_frame.columnconfigure(1, weight=1)

        # Excel başlık ve açıklama
        excel_title = tk.Label(excel_frame, text="Excel Dosyası", font=("Arial", 18), bg="#B0B0B0")
        excel_title.grid(row=0, column=0, columnspan=2, pady=(10,2), sticky="ew")
        
        excel_description = tk.Label(excel_frame, text="Bu bölümde excel dosyanızdaki ürünler masaüstünüze indirilir ve renk tahminleri yapılır. Tahmin edilen cevaplar yüklediğiniz excel dosyasına yazılır. Güncellenmiş excel dosyanız hazır olduğunda 'GÜNCELLENMİŞ EXCEL DOSYASINI AÇ' butonu aktif olacaktır.", font=("Arial", 8), wraplength=350 ,bg="#B0B0B0")
        excel_description.grid(row=1, column=0, columnspan=2, pady=(0,20), sticky="ew")
        
        # Excel seçimi ve klasör butonları yan yana
        excel_button = tk.Button(excel_frame, text="EXCEL DOSYASINI SEÇ", font=("Arial",10), wraplength=150, bg="#FFFFFF", command=self.select_excel)
        excel_button.grid(row=2, column=0, padx=10, pady=(15,2), sticky="ew")
        
        folder_button = tk.Button(excel_frame, text="KLASÖRÜ SEÇ", font=("Arial", 10), wraplength=200, bg="#FFFFFF", command=self.select_folder)
        folder_button.grid(row=2, column=1, padx=10, pady=(15,2), sticky="ew")
        
        # Seçilen dosya ve klasör etiketi
        self.excel_label = tk.Label(excel_frame, text="Excel dosyası seçilmedi", bg="#B0B0B0", font=("Arial", 10))
        self.excel_label.grid(row=3, column=0, padx=10, pady=(2,0), sticky="ew")
        
        self.folder_label = tk.Label(excel_frame, text="Görsellerin indirileceği klasör seçilmedi", wraplength=170, bg="#B0B0B0", font=("Arial", 10))
        self.folder_label.grid(row=3, column=1, padx=10, pady=(2,0), sticky="ew")
        
        run_button = tk.Button(excel_frame, text="ÇALIŞTIR", font=("Arial", 12), bg="#FFFFFF", command=self.run_process_thread)
        run_button.grid(row=4, column=0, columnspan=2, pady=(20,2), sticky="ew")
        
        self.run_label=tk.Label(excel_frame, text="...", bg="#B0B0B0", font=("Arial", 10))
        self.run_label.grid(row=5, column=0, columnspan=2, padx=10, pady=(2,0), sticky="ew")
        
        # Güncellenmiş Excel Dosyasını aç Butonu
        open_excel_button = tk.Button(excel_frame, text="GÜNCELLENMİŞ EXCEL DOSYASINI AÇ", font=("Arial", 12), bg="#FFFFFF", command=self.open_excel)
        open_excel_button.grid(row=6, column=0, columnspan=2, pady=(20,2), sticky="ew")
        
        # Görsel Dosyası Bölümü
        image_frame = tk.Frame(self.root, bg="#B0B0B0")
        image_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        image_frame.columnconfigure(0, weight=1)
        
        # Görsel başlık ve açıklama
        image_title = tk.Label(image_frame, text="Görsel", font=("Arial", 18), bg="#B0B0B0")
        image_title.pack(pady=(10,2), fill="x")
        
        image_description = tk.Label(image_frame, text="Bu bölümde yükleyeceğiniz görseldeki ürün analiz edilerek renk tahmini yapılır. Analiz tamamlandığında, yapay zekanın renk tahmini 'ÇALIŞTIR' butonun altında yazacaktır.", font=("Arial", 8), wraplength=350 ,bg="#B0B0B0")
        image_description.pack(pady=(0,20), fill="x")
        
        # Görsel ve çalıştır butonları
        image_button = tk.Button(image_frame, text="GÖRSELİ SEÇ", font=("Arial", 10), bg="#FFFFFF", command=self.select_image)
        image_button.pack(pady=(30,2), fill="x")
        
        self.image_label = tk.Label(image_frame, text="Görsel seçilmedi", bg="#B0B0B0", font=("Arial", 10))
        self.image_label.pack(pady=(2,0), fill="x")
        
        run_image_button = tk.Button(image_frame, text="ÇALIŞTIR", font=("Arial", 12), bg="#FFFFFF", command=self.run_image_process_thread)
        run_image_button.pack(pady=(35, 2), fill="x")

        self.run_image_label = tk.Label(image_frame, text=f"...", bg="#B0B0B0", font=("Arial", 10))
        self.run_image_label.pack(pady=(2, 0), fill="x")


    
    ai_excel_file_path = None
    def open_excel(self):
        open_excel(self.ai_excel_file_path)
        
    def select_excel(self):
        self.excel_file = select_excel_file(filedialog, self.excel_label, self.run_label)
    
    def select_image(self):
        self.image_file = select_image_file(filedialog, self.image_label, self.image_label, self.run_image_label)

    def select_folder(self):
        self.folder_of_path = select_folder(filedialog, self.folder_label, self.run_label)

    def run_process_thread(self):
        threading.Thread(target=self.run_process).start()

    def run_process(self):
        # Excel yükleme ve işleme süreci                                                                    # "Hayattan dışarı çıkıp soluklansam." diyorsun
        if not self.excel_file and not self.folder_of_path:                                                 # Bilmediğin onlarca şeyin peşinden koşuyorsun   
            messagebox.showerror("Hata!", "Excel dosyası ve klasör seçin.")                                 # Hatır için yaşamaktan fazlasını yapmalısın              
            self.ai_excel_file_path = None                                                                  # Yoksa gece yağınca kapkaranlık olacaksın
        elif not self.excel_file:
            messagebox.showerror("Hata!", "Bir excel dosyası seçin.")
            self.ai_excel_file_path = None
        elif not self.folder_of_path:
            messagebox.showerror("Hata!", "Bir klasör seçin.")
            self.ai_excel_file_path = None
            return
        else:
            df = load_workbook(filename=self.excel_file)
            sheet = df.active
            url_sutun_index = find_url_column(sheet)
            if url_sutun_index is None:
                messagebox.showerror("Hata!", "URL sütunu bulunamadı.")
                return
    
            # Yapay zeka modelinin tahminleri için sütun oluşturma
            yapay_zeka_sutun = url_sutun_index + 1
            sheet.insert_cols(yapay_zeka_sutun)  # Yeni sütun, URL sütunundan sonra ekleniyor
            column_letter = get_column_letter(url_sutun_index + 1)  # Yeni sütun harfi
            sheet[f'{column_letter}1'] = "Yapay Zeka Renk Tahminleri"
            self.run_label.config(text=f"Lütfen analiz tamamlanana kadar bekleyiniz...")
            messagebox.showinfo("Analizi Başladı!", "Yapay zeka analizi başladı, lütfen analizin tamamlanmasını bekleyiniz.")
        
            exit_loop = False  # Döngüden çıkmak için bir flag değişkeni oluşturun
            for cell in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=url_sutun_index, max_col=url_sutun_index):
                for url_cell in cell:
                    if url_cell.value is None:  # URL'nin boş olup olmadığını kontrol edin
                        exit_loop = True  # Döngüden çıkmak için flag
                        break
                    elif not url_cell.value.startswith(('http://', 'https://')):  # URL'nin geçerli bir formatta olup olmadığını kontrol edin
                        messagebox.showerror("Hata!", f"Geçersiz URL: {url_cell.value}")
                        continue  # Geçersiz URL'yi atla
                    else:
                        file_path = download_image(url_cell.value, self.folder_of_path)
                        if file_path:
                            prediction = predict_color(self.model, edit_image(file_path))
                            sheet.cell(row=url_cell.row, column=yapay_zeka_sutun).value = prediction
                        else:
                            messagebox.showerror("Hata!", f"{url_cell.value} URL'si indirilemedi.")
                if exit_loop:
                    break

            # Güncellenmiş dosyayı kaydetmek için yeni doysa adı oluşturacağız...
            excel_path = Path(self.excel_file)
            excel_base_name = excel_path.stem  # Eski .xlsx uzantısını kaldırır
            ai_file_name = f"{excel_base_name} Yapay Zeka Sonuçları.xlsx"
            # Dosya yolunu oluşturalım
            self.ai_excel_file_path = excel_path.parent / ai_file_name
            try:
                df.save(self.ai_excel_file_path)
                messagebox.showinfo("Başarılı!", f"Görseller indirildi ve analiz tamamlandı! Güncellenmiş dosya, {self.ai_excel_file_path} adresine kaydedildi.")
                self.run_label.config(text=f"Analiz tamamlanandı! Güncellenmiş dosyayı açabilirsiniz.")
            except Exception as e:
                print(f"Dosya kaydedilemedi: {e}")
                messagebox.showerror("Hata", f"Dosya kaydedilemedi: {e}")

    def run_image_process_thread(self):
        threading.Thread(target=self.run_image_process).start()
        
    def run_image_process(self):
        if not self.image_file:
            messagebox.showerror("Hata!", "Görseli seçin.")
            return
        prediction = predict_color(self.model, edit_image(self.image_file))
        self.run_image_label.config(text=f"Yapay Zeka Renk Tahmini: {prediction}")
