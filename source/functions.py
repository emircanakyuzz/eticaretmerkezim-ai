import os
import requests
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image
from tkinter import messagebox

def select_excel_file(filedialog, label, run_label):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        label.config(text=f"Seçilen dosya: {os.path.basename(file_path)}")
        return file_path
    else:
        label.config(text="Excel dosyası seçilmedi")
        run_label.config(text="...")
        return None
    
def select_image_file(filedialog, label, image_label, run_image_label):
    image_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
    if image_path:
        image_label.config(text=f"Seçilen görsel: {os.path.basename(image_path)}")
        return image_path
    else:
        image_path=None
        image_label.config(text="Görsel seçilmedi...")
        run_image_label.config(text= f"Yapay zekanın renk tahmini için bir görsel seçiniz.")
        return None

def select_folder(filedialog, label, run_label):
    folder_path = filedialog.askdirectory()
    if folder_path:
        label.config(text=f"Görsellerin indirileceği klasör: {os.path.basename(folder_path)}")
        return folder_path
    else:
        label.config(text="Klasör seçilmedi")
        run_label.config(text="...")
        return None

def open_excel(file_path):
        if file_path is None:
            messagebox.showerror("Hata!", f"Excel dosyası henüz güncellenmedi...")
        else:
            os.startfile(file_path)

def find_url_column(sheet):
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if str(cell_value).lower() == "url":
            return col
    return None

def edit_image(image_path):
    new_image = Image.open(image_path).resize((128, 128))
    image_array = np.array(new_image) / 255.0
    return np.expand_dims(image_array, axis=0)

def download_image(url, folder_of_path):
    file_path = os.path.join(folder_of_path, 'image.jpg')
    response = requests.get(url, verify=False)
    if response.status_code == 200:
        with open(file_path, 'wb') as file:
            file.write(response.content)
        return file_path
    else:
        return None

